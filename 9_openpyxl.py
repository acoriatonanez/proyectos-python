import openpyxl
libro = openpyxl.load_workbook("Dataset-Supertienda.xlsx")
hoja = libro["Ventas"]
total_venta = 0
total_ganancia = 0
for fila in hoja.iter_rows(min_row=2,values_only=True):
    if fila[16] is None: continue
    total_venta = total_venta + fila[16]
    total_ganancia = total_ganancia + fila[19]
margen = (total_ganancia/total_venta)*100
print(f"${total_venta:.2f}")
print(f"${total_ganancia:.2f}")
print(f"%{margen:.2f}")
libro_resumen = openpyxl.Workbook()
hoja_resumen = libro_resumen.active
hoja_resumen.title = "Resumen"
hoja_resumen.cell(row=1,column=1,value="total venta")
hoja_resumen.cell(row=2,column=1,value="total ganancia")
hoja_resumen.cell(row=3,column=1,value="margen")
hoja_resumen.cell(row=1,column=2,value=round(total_venta,2))
hoja_resumen.cell(row=2,column=2,value=round(total_ganancia,2))
hoja_resumen.cell(row=3,column=2,value=round(margen,2))
libro_resumen.save("libro_resumen.xlsx")